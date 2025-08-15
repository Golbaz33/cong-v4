# Fichier : utils/file_utils.py
# Version finale corrigée pour être compatible avec le threading de SQLite.

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from datetime import datetime

# On importe les classes nécessaires pour créer notre propre connexion
from db.database import DatabaseManager
from utils.config_loader import CONFIG
from utils.date_utils import format_date_for_display

def _perform_db_operation(db_path, operation_callback):
    """Fonction utilitaire pour gérer la connexion/déconnexion DB dans un thread."""
    db = DatabaseManager(db_path)
    if not db.connect():
        raise ConnectionError("Impossible de se connecter à la base de données depuis le thread.")
    try:
        # Exécute l'opération de lecture/écriture
        result = operation_callback(db)
        return result
    finally:
        # Assure que la connexion est toujours fermée
        db.close()

def export_agents_to_excel(db_path, save_path):
    """Exporte la liste des agents. Conçu pour être exécuté dans un thread."""
    def operation(db):
        agents = db.get_agents()
        if not agents:
            return "Aucun agent à exporter."
        
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Agents"
        headers = ["ID", "Nom", "Prénom", "PPR", "Grade", "Solde"]
        ws.append(headers)
        header_font = Font(bold=True)
        for cell in ws[1]: cell.font = header_font
        for agent in agents:
            ws.append([agent.id, agent.nom, agent.prenom, agent.ppr, agent.grade, agent.solde])
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value or "")) for cell in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
        wb.save(save_path)
        return f"Liste des agents exportée avec succès vers\n{save_path}"

    return _perform_db_operation(db_path, operation)

def export_all_conges_to_excel(db_path, save_path):
    """Exporte tous les congés. Conçu pour être exécuté dans un thread."""
    def operation(db):
        all_conges = db.get_conges()
        if not all_conges:
            return "Aucun congé à exporter."
            
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Tous les Congés"
        headers = ["Nom Agent", "Prénom Agent", "PPR Agent", "Type Congé", "Début", "Fin", "Jours Pris", "Statut", "Justification", "Intérimaire"]
        ws.append(headers)
        header_font = Font(bold=True)
        for cell in ws[1]: cell.font = header_font
        all_agents = {agent.id: agent for agent in db.get_agents()}
        for conge in all_conges:
            agent = all_agents.get(conge.agent_id)
            agent_nom, agent_prenom, agent_ppr = (agent.nom, agent.prenom, agent.ppr) if agent else ("Agent", "Supprimé", "")
            interim_info = ""
            if conge.interim_id:
                interim = all_agents.get(conge.interim_id)
                interim_info = f"{interim.nom} {interim.prenom}" if interim else "Agent Supprimé"
            row_data = [agent_nom, agent_prenom, agent_ppr, conge.type_conge, format_date_for_display(conge.date_debut), format_date_for_display(conge.date_fin), conge.jours_pris, conge.statut, conge.justif or "", interim_info]
            ws.append(row_data)
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_length = max(len(str(cell.value or "")) for cell in col_cells)
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2
        wb.save(save_path)
        return f"Tous les congés ont été exportés avec succès vers\n{save_path}"

    return _perform_db_operation(db_path, operation)

def import_agents_from_excel(db_path, source_path):
    """Importe des agents. Conçu pour être exécuté dans un thread."""
    def operation(db):
        errors = []; added_count, updated_count = 0, 0
        agent_import_headers = CONFIG['agent_import_headers']
        grades = CONFIG['ui']['grades']
        default_grade = grades[0] if grades else "Administrateur"
        default_solde = 22.0
        
        wb = openpyxl.load_workbook(source_path)
        ws = wb.active
        header = [str(cell.value or '').lower().strip() for cell in ws[1]]
        if not all(h in header for h in agent_import_headers):
            raise ValueError(f"Colonnes requises : {', '.join(agent_import_headers)}")

        col_map = {name: i for i, name in enumerate(header)}
        
        db.conn.execute('BEGIN TRANSACTION')
        try:
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if all(c is None for c in row): continue
                try:
                    nom = str(row[col_map['nom']] or '').strip(); prenom = str(row[col_map['prenom']] or '').strip()
                    if not nom or not prenom: raise ValueError("Nom et prénom obligatoires.")
                    ppr = str(row[col_map['ppr']] or '').strip() or f"{nom.upper()[:4]}_{datetime.now().strftime('%f')}"
                    grade = str(row[col_map['grade']] or '').strip() or default_grade
                    if grade not in grades: raise ValueError(f"Grade '{grade}' invalide.")
                    solde = float(str(row[col_map['solde']] or '').strip().replace(",", ".")) if row[col_map['solde']] is not None else default_solde
                    if solde < 0: raise ValueError("Solde négatif.")
                    
                    existing_agent = db.execute_query("SELECT id FROM agents WHERE ppr=?", (ppr,), fetch="one")
                    if existing_agent:
                        db.modifier_agent(existing_agent[0], nom, prenom, ppr, grade, solde); updated_count += 1
                    else:
                        db.ajouter_agent(nom, prenom, ppr, grade, solde); added_count += 1
                except Exception as ve:
                    errors.append(f"Ligne {i}: {ve}")
            
            if errors:
                db.conn.rollback()
                raise Exception("Importation annulée. Erreurs:\n" + "\n".join(errors[:5]))
            else:
                db.conn.commit()
                return f"Importation réussie !\n\n- Agents ajoutés: {added_count}\n- Agents mis à jour: {updated_count}"
        except Exception as e:
            db.conn.rollback()
            raise e # Propage l'erreur pour qu'elle soit attrapée par le thread handler

    return _perform_db_operation(db_path, operation)